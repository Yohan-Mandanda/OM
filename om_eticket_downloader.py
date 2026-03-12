from __future__ import annotations

import argparse
import csv
import re
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, List, Optional
from urllib.parse import urljoin, urlparse, unquote

from playwright.sync_api import BrowserContext, Page, TimeoutError, sync_playwright


BASE_URL = "https://billetterie.om.fr/fr/"


@dataclass
class Account:
    email: str
    password: str


@dataclass
class DownloadResult:
    email: str
    match_name: str
    saved_files: List[Path]
    success: bool
    error: Optional[str] = None


def _normalize(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value or "")
    ascii_value = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", ascii_value).strip().lower()


def _sanitize_segment(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", value).strip("_") or "account"


def _parse_match_queries(value: str) -> List[str]:
    parts = [p.strip() for p in re.split(r"[,\n;]+", value or "") if p.strip()]
    unique: List[str] = []
    seen = set()
    for part in parts:
        key = _normalize(part)
        if key and key not in seen:
            unique.append(part)
            seen.add(key)
    return unique


def _resolve_header(headers: List[str], candidates: Iterable[str]) -> Optional[str]:
    normalized_headers = {_normalize(h): h for h in headers if h}
    for candidate in candidates:
        if candidate in normalized_headers:
            return normalized_headers[candidate]
    return None


def load_accounts_from_file(file_path: Path) -> List[Account]:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return _rows_to_accounts(list(reader), reader.fieldnames or [])

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for Excel files. Install dependencies from requirements.txt."
            ) from exc

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows = []
        for row in rows[1:]:
            row_dict = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else ""
                row_dict[header] = "" if value is None else str(value).strip()
            data_rows.append(row_dict)
        return _rows_to_accounts(data_rows, headers)

    raise ValueError("Unsupported account file format. Use CSV or XLSX.")


def _rows_to_accounts(rows: List[dict], headers: List[str]) -> List[Account]:
    email_header = _resolve_header(
        list(headers),
        ["email", "adresse email", "adresseemail", "mail", "e-mail", "login"],
    )
    password_header = _resolve_header(
        list(headers),
        ["mot de passe", "motdepasse", "password", "pass", "mdp"],
    )

    if not email_header or not password_header:
        raise ValueError(
            "Missing columns. Expected email and password columns "
            "(e.g. 'email' + 'mot de passe' or 'password')."
        )

    accounts: List[Account] = []
    for row in rows:
        email = (row.get(email_header) or "").strip()
        password = (row.get(password_header) or "").strip()
        if email and password:
            accounts.append(Account(email=email, password=password))
    return accounts


def _emit(progress_cb: Optional[Callable[[str], None]], message: str) -> None:
    if progress_cb:
        progress_cb(message)


def _wait_between_actions(page: Page, step_wait_seconds: int) -> None:
    if step_wait_seconds > 0:
        page.wait_for_timeout(step_wait_seconds * 1000)


def _accept_cookie_banner(page: Page) -> None:
    cookie_labels = [
        "Tout accepter",
        "Accepter",
        "J'accepte",
        "Autoriser",
        "Accept all",
    ]
    for label in cookie_labels:
        try:
            button = page.get_by_role("button", name=re.compile(label, re.IGNORECASE))
            if button.count() > 0:
                button.first.click(timeout=2000)
                page.wait_for_timeout(500)
                return
        except Exception:
            continue


def _wait_for_login_state(page: Page, max_wait_seconds: int) -> None:
    end_time = time.time() + max_wait_seconds
    while time.time() < end_time:
        has_connect_text = (
            page.locator("#om-top-bar")
            .get_by_text(re.compile(r"se\s*connecter", re.IGNORECASE))
            .count()
            > 0
        )
        has_login_form = page.locator("#popup-login-login-form").count() > 0 and page.locator(
            "#popup-login-login-form"
        ).first.is_visible()
        if not has_connect_text and not has_login_form:
            return
        page.wait_for_timeout(1000)
    raise RuntimeError(
        "Login did not complete in time. Captcha may require manual solve in headed mode."
    )


def _open_profile_menu(page: Page) -> None:
    profile_button = page.locator("#om-top-bar > div > ul > li > a").first
    profile_button.wait_for(state="visible", timeout=15000)
    profile_button.click()


def _go_to_billetterie_space(page: Page) -> None:
    _open_profile_menu(page)
    billetterie_item = page.get_by_text(
        re.compile(r"mon espace billetterie", re.IGNORECASE), exact=False
    ).first
    billetterie_item.wait_for(state="visible", timeout=15000)
    billetterie_item.click()
    page.wait_for_load_state("domcontentloaded")
    page.locator("ul.manageEvtCardList").first.wait_for(timeout=20000)


def _event_cards_locator(page: Page):
    selectors = [
        "ul.manageEvtCardList li a",
        "ul.u-flex.manageEvtCardList > li > a",
        "li a:has(.manageEvtCardTitle)",
        "a:has(.manageEvtCardEnd)",
    ]
    for selector in selectors:
        locator = page.locator(selector)
        if locator.count() > 0:
            return locator
    return page.locator("ul.manageEvtCardList li a")


def _extract_card_title(card) -> str:
    title_selectors = [
        ".manageEvtCardTitle span",
        ".manageEvtCardTitle",
        "strong.manageEvtCardTitle",
    ]

    for selector in title_selectors:
        locator = card.locator(selector)
        if locator.count() == 0:
            continue
        try:
            raw = locator.first.inner_text(timeout=3000).strip()
        except Exception:
            continue
        if not raw:
            continue

        lines = [re.sub(r"\s+", " ", line).strip() for line in raw.splitlines()]
        lines = [line for line in lines if line]
        if lines:
            return lines[-1]

    # Last-resort extraction from the full card text.
    try:
        card_text = card.inner_text(timeout=3000)
    except Exception:
        return ""

    ignored_tokens = ("ligue 1", "journee", "orange velodrome", "week-end")
    lines = [re.sub(r"\s+", " ", line).strip() for line in card_text.splitlines()]
    candidates = [
        line
        for line in lines
        if line
        and any(ch.isalpha() for ch in line)
        and not any(token in _normalize(line) for token in ignored_tokens)
        and not re.search(r"\d{1,2}\s+\w+\s+\d{4}", _normalize(line))
    ]
    return candidates[-1] if candidates else ""


def _choose_match(page: Page, match_query: str) -> str:
    cards = _event_cards_locator(page)
    count = cards.count()
    if count == 0:
        raise RuntimeError("No match cards found in 'Mon espace billetterie'.")

    normalized_query = _normalize(match_query)
    found_card = None
    available = []

    for idx in range(count):
        card = cards.nth(idx)
        title = _extract_card_title(card)
        if not title:
            continue
        available.append(title)
        normalized_title = _normalize(title)
        if normalized_query in normalized_title or normalized_title in normalized_query:
            found_card = card
            chosen_title = title
            break

    if not found_card:
        raise RuntimeError(
            f"Match '{match_query}' not found. Available matches: {', '.join(available)}"
        )

    found_card.click()
    page.wait_for_load_state("domcontentloaded")
    return chosen_title


def _filename_from_headers(headers: dict, fallback_url: str, default_name: str) -> str:
    content_disposition = headers.get("content-disposition", "")
    if content_disposition:
        # RFC 6266 / 5987 handling for common filename formats.
        match_utf8 = re.search(r"filename\*=UTF-8''([^;]+)", content_disposition, re.IGNORECASE)
        if match_utf8:
            return unquote(match_utf8.group(1).strip('"'))
        match_basic = re.search(r'filename="?([^";]+)"?', content_disposition, re.IGNORECASE)
        if match_basic:
            return match_basic.group(1)

    parsed = urlparse(fallback_url)
    basename = Path(unquote(parsed.path)).name
    if basename:
        return basename
    return default_name


def _unique_file_path(target_path: Path) -> Path:
    if not target_path.exists():
        return target_path

    stem = target_path.stem
    suffix = target_path.suffix
    for idx in range(1, 500):
        candidate = target_path.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Could not create unique file name for {target_path.name}")


def _download_via_request(context: BrowserContext, url: str, destination_dir: Path) -> Path:
    response = context.request.get(url, timeout=45000)
    if not response.ok:
        raise RuntimeError(f"Download failed with HTTP {response.status} for URL {url}")

    filename = _filename_from_headers(response.headers, url, "eticket.pdf")
    target = _unique_file_path(destination_dir / filename)
    target.write_bytes(response.body())
    return target


def _download_ticket_buttons(page: Page, context: BrowserContext, destination_dir: Path) -> List[Path]:
    destination_dir.mkdir(parents=True, exist_ok=True)
    downloaded: List[Path] = []

    buttons = page.locator(
        "a:has(span.ctaFullLabel:has-text('Télécharger')), button:has-text('Télécharger')"
    )
    count = buttons.count()
    if count == 0:
        raise RuntimeError("No 'Télécharger' button found for this match.")

    for idx in range(count):
        button = buttons.nth(idx)
        href = button.get_attribute("href")
        pages_before = {id(p): p for p in context.pages}

        try:
            with page.expect_download(timeout=12000) as dl_info:
                button.click()
            download = dl_info.value
            target = _unique_file_path(destination_dir / download.suggested_filename)
            download.save_as(str(target))
            downloaded.append(target)
            continue
        except TimeoutError:
            pass

        # Fallback 1: direct authenticated request from href.
        if href and not href.startswith("#") and not href.lower().startswith("javascript:"):
            absolute_url = urljoin(page.url, href)
            downloaded.append(_download_via_request(context, absolute_url, destination_dir))
            continue

        # Fallback 2: popup opens with PDF URL.
        popup = None
        for _ in range(20):
            new_pages = [p for pid, p in {id(p): p for p in context.pages}.items() if pid not in pages_before]
            if new_pages:
                popup = new_pages[0]
                break
            page.wait_for_timeout(300)

        if popup:
            popup.wait_for_load_state("domcontentloaded", timeout=15000)
            if popup.url and popup.url != "about:blank":
                downloaded.append(_download_via_request(context, popup.url, destination_dir))
            popup.close()
            continue

        raise RuntimeError(
            f"Could not capture download for ticket button #{idx + 1}. "
            "Try headed mode and verify selectors."
        )

    return downloaded


def _logout(page: Page) -> None:
    _open_profile_menu(page)
    logout_item = page.get_by_text(re.compile(r"d[ée]connexion", re.IGNORECASE), exact=False).first
    logout_item.wait_for(state="visible", timeout=10000)
    logout_item.click()
    page.wait_for_load_state("domcontentloaded")


def _process_account(
    context: BrowserContext,
    account: Account,
    match_name: str,
    destination_root: Path,
    login_wait_seconds: int,
    step_wait_seconds: int,
    progress_cb: Optional[Callable[[str], None]],
) -> DownloadResult:
    page = context.new_page()
    account_dir = destination_root / _sanitize_segment(account.email)
    saved_files: List[Path] = []
    match_queries = _parse_match_queries(match_name)
    matched_names: List[str] = []

    try:
        if not match_queries:
            raise RuntimeError("No match name provided.")

        _emit(progress_cb, f"[{account.email}] Opening site...")
        page.goto(BASE_URL, wait_until="domcontentloaded", timeout=60000)
        _accept_cookie_banner(page)
        _wait_between_actions(page, step_wait_seconds)

        _emit(progress_cb, f"[{account.email}] Logging in...")
        page.get_by_text(re.compile(r"se\s*connecter", re.IGNORECASE), exact=False).first.click(
            timeout=15000
        )
        _wait_between_actions(page, step_wait_seconds)
        page.locator("input[name='popup-login-email']").fill(account.email)
        _wait_between_actions(page, step_wait_seconds)
        page.locator("input[name='popup-login-password']").fill(account.password)
        _wait_between_actions(page, step_wait_seconds)
        page.locator("#popup-login-login-form button[type='submit']").click()
        _wait_between_actions(page, step_wait_seconds)

        _wait_for_login_state(page, login_wait_seconds)
        _emit(progress_cb, f"[{account.email}] Opening billetterie space...")
        _go_to_billetterie_space(page)
        _wait_between_actions(page, step_wait_seconds)

        for idx, query in enumerate(match_queries):
            chosen_match = _choose_match(page, query)
            matched_names.append(chosen_match)
            _emit(progress_cb, f"[{account.email}] Match selected: {chosen_match}")
            _wait_between_actions(page, step_wait_seconds)

            new_files = _download_ticket_buttons(page, context, account_dir)
            saved_files.extend(new_files)
            _emit(
                progress_cb,
                f"[{account.email}] Downloaded {len(new_files)} file(s) for '{chosen_match}'.",
            )
            _wait_between_actions(page, step_wait_seconds)

            # If more matches are requested, return to list view and continue.
            if idx < len(match_queries) - 1:
                page.goto(BASE_URL, wait_until="domcontentloaded", timeout=60000)
                _wait_between_actions(page, step_wait_seconds)
                _go_to_billetterie_space(page)
                _wait_between_actions(page, step_wait_seconds)

        _logout(page)
        _emit(progress_cb, f"[{account.email}] Logged out.")
        return DownloadResult(
            email=account.email,
            match_name=", ".join(matched_names) if matched_names else match_name,
            saved_files=saved_files,
            success=True,
        )
    except Exception as exc:
        return DownloadResult(
            email=account.email,
            match_name=match_name,
            saved_files=saved_files,
            success=False,
            error=str(exc),
        )
    finally:
        page.close()


def run_eticket_downloads(
    accounts: List[Account],
    match_name: str,
    output_dir: Path,
    headless: bool = False,
    login_wait_seconds: int = 120,
    step_wait_seconds: int = 10,
    slow_mo_ms: int = 0,
    progress_cb: Optional[Callable[[str], None]] = None,
) -> List[DownloadResult]:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    results: List[DownloadResult] = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms)
        try:
            for account in accounts:
                context = browser.new_context(accept_downloads=True)
                try:
                    result = _process_account(
                        context=context,
                        account=account,
                        match_name=match_name,
                        destination_root=output_path,
                        login_wait_seconds=login_wait_seconds,
                        step_wait_seconds=step_wait_seconds,
                        progress_cb=progress_cb,
                    )
                    results.append(result)
                finally:
                    context.close()
        finally:
            browser.close()
    return results


def _build_cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Download OM e-tickets for multiple accounts.")
    parser.add_argument(
        "--accounts-file",
        required=True,
        help="Path to CSV or XLSX file containing email/password columns.",
    )
    parser.add_argument("--match", required=True, help="Match filter (e.g. 'Lille', 'Auxerre').")
    parser.add_argument(
        "--output-dir",
        default="downloads",
        help="Directory where e-ticket files are saved.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run browser in headless mode. Use headed mode when captcha appears.",
    )
    parser.add_argument(
        "--login-wait-seconds",
        type=int,
        default=120,
        help="How long to wait for login completion/captcha solve.",
    )
    parser.add_argument(
        "--step-wait-seconds",
        type=int,
        default=10,
        help="Pause between key UI actions (helps with slow page transitions).",
    )
    parser.add_argument(
        "--slow-mo-ms",
        type=int,
        default=0,
        help="Slow down browser actions (ms) for debugging.",
    )
    return parser


def main() -> int:
    parser = _build_cli()
    args = parser.parse_args()

    accounts = load_accounts_from_file(Path(args.accounts_file))
    if not accounts:
        print("No accounts found in the input file.")
        return 1

    print(f"Loaded {len(accounts)} account(s). Starting automation...")
    results = run_eticket_downloads(
        accounts=accounts,
        match_name=args.match,
        output_dir=Path(args.output_dir),
        headless=args.headless,
        login_wait_seconds=args.login_wait_seconds,
        step_wait_seconds=args.step_wait_seconds,
        slow_mo_ms=args.slow_mo_ms,
        progress_cb=print,
    )

    failures = [r for r in results if not r.success]
    print("\nSummary")
    print("-" * 40)
    for result in results:
        if result.success:
            print(f"[OK] {result.email} -> {len(result.saved_files)} file(s)")
        else:
            print(f"[FAIL] {result.email} -> {result.error}")
    print("-" * 40)
    print(f"Success: {len(results) - len(failures)}/{len(results)}")

    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
